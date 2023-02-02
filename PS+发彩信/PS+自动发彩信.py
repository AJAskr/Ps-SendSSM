import openpyxl
import hashlib
import ID
from PySide2.QtGui import QIcon
from PySide2.QtWidgets import QApplication, QMessageBox, QFileDialog, QTableWidgetItem, QLineEdit, QInputDialog
from PySide2.QtUiTools import QUiLoader
from configobj import ConfigObj
import subprocess
import os
import pyperclip
import keyboard
import companySereen
from aotuJob import Job


class Stats:

    def __init__(self):
        # 从文件中加载UI定义
        # 从 UI 定义中动态 创建一个相应的窗口对象
        # 注意：里面的控件对象也成为窗口对象的属性了
        # 比如 self.ui.button , self.ui.textEdit
        self.ui = QUiLoader().load("main.ui")
        # 获取ID并显示 然后检查是否激活
        self.disk_sn_list = ID.Hardware().get_disk_sn()
        self.md5_str = hashlib.md5(self.disk_sn_list[0].encode("utf-8")).hexdigest()
        self.ui.label_16.setText(self.md5_str)
        self.jihuo = self.isJihuo()
        # self.isjihuo()
        # 复制ID
        self.ui.pushButton_10.clicked.connect(self.copyID)
        # 点击激活
        self.ui.pushButton_2.clicked.connect(self.shuoming)
        # 选择文件
        self.ui.pushButton_9.clicked.connect(self.chooseFile)
        # 刷新配置
        self.updatePz()
        # 刷新坐标
        self.updatezb()
        # 编辑排版
        self.ui.pushButton_17.clicked.connect(self.oppaiban)
        # 保存配置
        self.ui.pushButton_3.clicked.connect(self.savePz)
        # 对表格内公司进行剔除多余字符
        self.ui.pushButton_16.clicked.connect(self.tichu)
        # 刷新
        self.ui.pushButton_4.clicked.connect(self.updatePz)
        # 重置
        self.ui.pushButton_15.clicked.connect(self.chongzhi)
        # 打开坐标工具
        self.ui.pushButton_5.clicked.connect(self.openzb)
        # 打开投屏工具
        self.ui.pushButton_6.clicked.connect(self.opentp)
        # 刷新坐标点击事件
        self.ui.pushButton_7.clicked.connect(self.updatezb)
        # 坐标被改动事件
        self.ui.tableWidget.cellChanged.connect(self.chagezb)
        # 插入一条数据到表格
        self.ui.pushButton_13.clicked.connect(self.insertzb)
        # 删除出表格最后一条数据
        self.ui.pushButton_14.clicked.connect(self.deletezb)
        # 插入一条数据到指定位置下方
        self.ui.pushButton_18.clicked.connect(self.insertByIndex)
        # 删除指定行
        self.ui.pushButton_19.clicked.connect(self.deleteBById)
        # 使用教程
        self.ui.pushButton_8.clicked.connect(self.jiaochen)
        # 需要公司
        self.ui.pushButton_20.clicked.connect(self.needCompany1)
        # 不需要公司
        self.ui.pushButton_21.clicked.connect(self.needCompany2)
        # 公司初始设定
        self.ineedcp()
        # 使用脚本
        self.okRun()

    # ##复制ID事件
    def copyID(self):
        ID = self.ui.label_16.text()
        pyperclip.copy(ID)
        QMessageBox.about(self.ui, "提示", "复制成功")

    # ##激活事件
    def isJihuo(self):
        mm = 'Aja' + self.md5_str + '935592713'
        mm = hashlib.md5(mm.encode("utf-8")).hexdigest()
        config1 = ConfigObj("./config/表格配置.ini", encoding="UTF8")
        value = config1['key']
        if str(value['1']) == mm:
            self.ui.label_3.setText("已激活")
            return True
        return False

    # ##激活事件
    def shuoming(self):
        mm = 'Aja' + self.md5_str + '935592713'
        mm = hashlib.md5(mm.encode("utf-8")).hexdigest()
        title, OK = QInputDialog.getText(self.ui, "输入目录名称", "名称:", QLineEdit.Normal, "1")
        if OK:
            if mm == title:
                config1 = ConfigObj("./config/表格配置.ini", encoding="UTF8")
                value = config1['key']
                value['1'] = title
                config1.write()
                self.ui.label_3.setText("已激活")

    # ##选择文件
    def chooseFile(self):
        filePath = QFileDialog.getOpenFileName(self.ui, "选择表格文件", "C:\\", "表格类型(*.xlsx *.xls)")
        # 修改路径输入框的值
        self.ui.textEdit.setPlainText(str(filePath[0]))

    # ##表格配置相关
    # ##获取表格配置的文本并修改在界面的控件上 --- 表格配置.ini
    def updatePz(self):
        config1 = ConfigObj("./config/表格配置.ini", encoding="UTF8")
        value = config1['BG']
        self.ui.textEdit.setPlainText(value['1'])
        self.ui.textEdit_2.setPlainText(value['2'])
        self.ui.textEdit_3.setPlainText(value['3'])
        self.ui.textEdit_4.setPlainText(value['4'])
        self.ui.textEdit_5.setPlainText(value['5'])
        self.ui.textEdit_7.setPlainText(value['7'])
        self.ui.textEdit_8.setPlainText(value['8'])
        self.ui.textEdit_6.setPlainText(value['9'])
        self.ui.label_13.setText(value['6'] + " 行")
        self.ui.textEdit_9.setPlainText(value['10'])

    # ##重置
    def chongzhi(self):
        config1 = ConfigObj("./config/表格配置.ini", encoding="UTF8")
        value = config1['BG']
        value['6'] = 0
        config1.write()
        self.updatePz()

    # ##保存表格配置.ini
    def savePz(self):
        config1 = ConfigObj("./config/表格配置.ini", encoding="UTF8")
        value = config1['BG']
        value['1'] = self.ui.textEdit.toPlainText()
        value['2'] = self.ui.textEdit_2.toPlainText()
        value['3'] = self.ui.textEdit_3.toPlainText()
        value['4'] = self.ui.textEdit_4.toPlainText()
        value['5'] = self.ui.textEdit_5.toPlainText()
        value['7'] = self.ui.textEdit_7.toPlainText()
        value['8'] = self.ui.textEdit_8.toPlainText()
        value['9'] = self.ui.textEdit_6.toPlainText()
        value['10'] = self.ui.textEdit_9.toPlainText()
        config1.write()
        QMessageBox.about(self.ui, "提示", "保存成功！      ")

    # ##剔除表格内公司名称多余字符事件
    def tichu(self):
        config1 = ConfigObj("./config/表格配置.ini", encoding="UTF8")
        value = config1['BG']
        if os.path.exists(value['1']):
            bgpath = str(value['1'])
            gslie = int(value['10'])
            kaishi = int(value['4'])
            jieshu = int((value['5']))
            try:
                workbook = openpyxl.load_workbook(bgpath)
                sheet = workbook.worksheets[0]
                for i in range(kaishi, jieshu + 1):
                    pname = sheet.cell(row=i, column=gslie).value
                    sr = companySereen.down(pname)
                    sheet.cell(row=i, column=gslie).value = sr
                workbook.save(bgpath)
            except:
                QMessageBox.about(self.ui, "提示", "文件处理异常      ")
            QMessageBox.about(self.ui, "提示", "文件处理成功      ")
        else:
            QMessageBox.about(self.ui, "提示", "找不到目标文件      ")

    # ##坐标配置相关
    # ##刷新坐标到表格的事件
    def updatezb(self):
        # 刷新前先清空不然会重复插入
        self.ui.tableWidget.setRowCount(0)
        config1 = ConfigObj("./config/坐标配置.ini", encoding="UTF8")
        value = config1['len']
        all = int(value['all'])
        # 工具配置文件创建多少行
        for i in range(0, all):
            self.ui.tableWidget.insertRow(i)
        # 读取配置文件的内容设置到表格内
        value2 = config1['remarks']
        value3 = config1['X']
        value4 = config1['Y']
        value5 = config1['event']
        value6 = config1['sleep']
        value7 = config1['copy']
        for i in range(0, all):
            index = str(i + 1)
            self.ui.tableWidget.setItem(i, 0, QTableWidgetItem(value2[index]))
            self.ui.tableWidget.setItem(i, 1, QTableWidgetItem(value3[index]))
            self.ui.tableWidget.setItem(i, 2, QTableWidgetItem(value4[index]))
            self.ui.tableWidget.setItem(i, 3, QTableWidgetItem(value5[index]))
            self.ui.tableWidget.setItem(i, 4, QTableWidgetItem(value6[index]))
            self.ui.tableWidget.setItem(i, 5, QTableWidgetItem(value7[index]))

    # ##把表格保存到坐标配置文件 只要改动就保存
    def chagezb(self, row, column):
        config1 = ConfigObj("./config/坐标配置.ini", encoding="UTF8")
        value = None
        if column == 0: value = config1['remarks']
        if column == 1: value = config1['X']
        if column == 2: value = config1['Y']
        if column == 3: value = config1['event']
        if column == 4: value = config1['sleep']
        if column == 5: value = config1['copy']
        index = str(row + 1)
        value[index] = self.ui.tableWidget.item(row, column).text()
        config1.write()

    # ##插入一行数据
    def insertzb(self):
        config1 = ConfigObj("./config/坐标配置.ini", encoding="UTF8")
        value1 = config1['len']
        value2 = config1['remarks']
        value3 = config1['X']
        value4 = config1['Y']
        value5 = config1['event']
        value6 = config1['sleep']
        value7 = config1['copy']
        index = str(int(value1['all']) + 1)
        value1['all'] = int(index)
        value2[index] = '描述'
        value3[index] = 100
        value4[index] = 200
        value5[index] = '鼠标移动到'
        value6[index] = 1
        value7[index] = '无'
        config1.write()
        self.updatezb()

    # ##删除一行数据
    def deletezb(self):
        if self.ui.tableWidget.rowCount() != 0:
            config1 = ConfigObj("./config/坐标配置.ini", encoding="UTF8")
            value1 = config1['len']
            index = str(value1['all'])
            value1['all'] = int(index) - 1
            del config1['remarks'][index]
            del config1['X'][index]
            del config1['Y'][index]
            del config1['event'][index]
            del config1['sleep'][index]
            del config1['copy'][index]
            config1.write()
            self.updatezb()
        else:
            QMessageBox.about(self.ui, "提示", "已经不能再删了      ")

    # ##指定位置插入
    def insertByIndex(self):
        title, ok = QInputDialog.getInt(self.ui, "在哪个位置下方插入", "位置序号:", QLineEdit.Normal, 1)
        if ok:
            try:
                config1 = ConfigObj("./config/坐标配置.ini", encoding="UTF8")
                value1 = config1['len']['all']
                value2 = config1['remarks']
                value3 = config1['X']
                value4 = config1['Y']
                value5 = config1['event']
                value6 = config1['sleep']
                value7 = config1['copy']
                index = int(title)
                if 1 <= index <= int(value1):
                    if index == int(value1):
                        self.insertzb()
                        return
                    index2 = str(int(value1) + 1)
                    value2[index2] = None
                    value3[index2] = None
                    value4[index2] = None
                    value5[index2] = None
                    value6[index2] = None
                    value7[index2] = None
                    config1['len']['all'] = index2
                    config1.write()
                    index3 = int(index2)
                    while index3 > index:
                        value2[str(index3)] = value2[str(index3 - 1)]
                        value3[str(index3)] = value3[str(index3 - 1)]
                        value4[str(index3)] = value4[str(index3 - 1)]
                        value5[str(index3)] = value5[str(index3 - 1)]
                        value6[str(index3)] = value6[str(index3 - 1)]
                        value7[str(index3)] = value7[str(index3 - 1)]
                        index3 -= 1
                    value2[str(index + 1)] = "插入"
                    value3[str(index + 1)] = "100"
                    value4[str(index + 1)] = "200"
                    value5[str(index + 1)] = "鼠标左键单击"
                    value6[str(index + 1)] = "1"
                    value7[str(index + 1)] = "无"
                    config1.write()
                    self.updatezb()
                else:
                    QMessageBox.about(self.ui, "提示", "输入的序号可能有误")
                    return
            except ValueError:
                QMessageBox.about(self.ui, "提示", "输入的序号可能有误")
                return

    # ##指定位置删除
    def deleteBById(self):
        title, ok = QInputDialog.getInt(self.ui, "删除哪一行数据？", "位置序号:", QLineEdit.Normal, 1)
        if ok:
            config1 = ConfigObj("./config/坐标配置.ini", encoding="UTF8")
            cfgLen = int(config1['len']['all'])
            if 1 <= int(title) <= cfgLen:
                if int(title) == cfgLen:
                    self.deletezb()
                    return
                index = 0
                for i in range(1, cfgLen):
                    if i == int(title):
                        index = 1
                    config1['remarks'][str(i)] = config1['remarks'][str(i+index)]
                    config1['X'][str(i)] = config1['X'][str(i+index)]
                    config1['Y'][str(i)] = config1['Y'][str(i+index)]
                    config1['event'][str(i)] = config1['event'][str(i+index)]
                    config1['sleep'][str(i)] = config1['sleep'][str(i+index)]
                    config1['copy'][str(i)] = config1['copy'][str(i+index)]
                del config1['remarks'][str(cfgLen)]
                del config1['X'][str(cfgLen)]
                del config1['Y'][str(cfgLen)]
                del config1['event'][str(cfgLen)]
                del config1['sleep'][str(cfgLen)]
                del config1['copy'][str(cfgLen)]
                config1['len']['all'] = cfgLen-1
                config1.write()
                self.updatezb()
            else:
                QMessageBox.about(self.ui, "提示", "输入的序号可能有误")
                return

    # ##打开坐标获取工具的事件
    def openzb(self):
        if os.path.exists("./getXY/获取鼠标坐标.exe"):
            subprocess.Popen("./getXY/获取鼠标坐标.exe")
        else:
            QMessageBox.about(self.ui, "提示", "工具被删除或者移动到某个位置      ")

    # ##打开投屏工具的事件
    def opentp(self):
        if os.path.exists("./tp/QtScrcpy-win-x64-v2.1.2"):
            path = os.path.abspath("./tp/QtScrcpy-win-x64-v2.1.2")
            os.startfile(path)
        else:
            QMessageBox.about(self.ui, "提示", "工具被删除或者移动到某个位置      ")

    # ##打开使用教程
    def jiaochen(self):
        try:
            path = os.path.abspath("./config/使用教程.docx")
            os.startfile(path)
        except:
            QMessageBox.about(self.ui, "提示", "没有安装word查看工具或者文件被删除      ")

    # ##打开排版文本
    def oppaiban(self):
        config1 = ConfigObj("./config/表格配置.ini", encoding="UTF8")
        value1 = config1['BG2']
        path = value1['2']
        os.startfile(path)

    # ##需要公司
    def needCompany1(self):
        config1 = ConfigObj("./config/表格配置.ini", encoding="UTF8")
        value1 = config1['BG2']
        value1['1'] = "是"
        config1.write()
        QMessageBox.about(self.ui, "提示", "设置成功      ")

    # ##不需要公司
    def needCompany2(self):
        config1 = ConfigObj("./config/表格配置.ini", encoding="UTF8")
        value1 = config1['BG2']
        value1['1'] = "否"
        config1.write()
        QMessageBox.about(self.ui, "提示", "设置成功      ")

    ##一开始就需要公司
    def ineedcp(self):
        config1 = ConfigObj("./config/表格配置.ini", encoding="UTF8")
        value1 = config1['BG2']
        value1['2'] = os.path.abspath('./config/排版.txt')
        config1.write()

    # ##执行脚本任务
    def runJob(self):
        if self.jihuo:
            a = Job()
            a.start()
            keyboard.add_hotkey('f12', a.stop)
        else:
            QMessageBox.about(self.ui, "提示", "您的设备没有权限~      ")

    # 按下f10开始脚本
    def okRun(self):
        keyboard.add_hotkey('f10', self.runJob)


app = QApplication([])
app.setWindowIcon(QIcon("icon.png"))
stats = Stats()
stats.ui.show()
app.exec_()
