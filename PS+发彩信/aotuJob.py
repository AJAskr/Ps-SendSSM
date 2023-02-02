import os

import openpyxl
import pyautogui
import threading
import time
import inspect
import ctypes
import pyperclip
from configobj import ConfigObj


def _async_raise(tid, exctype):
    """raises the exception, performs cleanup if needed"""

    tid = ctypes.c_long(tid)

    if not inspect.isclass(exctype):
        exctype = type(exctype)

    res = ctypes.pythonapi.PyThreadState_SetAsyncExc(tid, ctypes.py_object(exctype))

    if res == 0:

        raise ValueError("invalid thread id")

    elif res != 1:

        # """if it returns a number greater than one, you're in trouble,

        # and you should call it again with exc=NULL to revert the effect"""

        ctypes.pythonapi.PyThreadState_SetAsyncExc(tid, None)

        raise SystemError("PyThreadState_SetAsyncExc failed")


def stop_thread(thread):
    _async_raise(thread.ident, SystemExit)


class Job(threading.Thread):

    def __init__(self, *args, **kwargs):
        super(Job, self).__init__(*args, **kwargs)
        # 用于暂停线程的标识
        self.__flag = threading.Event()
        self.__flag.set()  # 设置为True
        # 用于停止线程的标识
        self.__running = threading.Event()
        self.__running.set()  # 将running设置为True

    def run(self):
        config1 = ConfigObj("./config/表格配置.ini", encoding="UTF8")
        config2 = ConfigObj("./config/坐标配置.ini", encoding="UTF8")
        zblen = config2['len']['all']
        value = config1['BG']
        bgpath = value['1']
        if not os.path.exists(bgpath):
            pyautogui.alert(text='没有找到表格文件~', title='警告', button='OK')
            self.stop()
        farenLie = value['2']
        sjhLie = value['3']
        kaishi = value['4']
        jieshu = value['5']
        xingxi = str(value['7']) + str(value['8'])
        kongge = ' ' * int(value['9'])
        gongsi = value['10']
        needgongsi = str(config1['BG2']['1'])
        pbpath = str(config1['BG2']['2'])
        workbook = openpyxl.load_workbook(bgpath)
        sheet = workbook.worksheets[0]
        i = int(kaishi)
        while self.__running.isSet():
            if i <= int(jieshu):
                farenName = str(sheet.cell(row=i, column=int(farenLie)).value)
                phoneNum = str(sheet.cell(row=i, column=int(sjhLie)).value)
                companyName = str(sheet.cell(row=i, column=int(gongsi)).value)
                resPaiban = paiban(pbpath,companyName)
                if len(farenName) == 2:
                    farenName = farenName[0] + kongge + farenName[1]
                for j in range(0, int(zblen)):
                    index = str(j + 1)
                    x = config2['X'][index]
                    y = config2['Y'][index]
                    event = config2['event'][index]
                    sleepNum = config2['sleep'][index]
                    copytext = str(config2['copy'][index])
                    if '>法人的姓+' in copytext:
                        wb = farenName[0]+copytext.split('+')[1]
                        pyperclip.copy(wb)
                    elif copytext == '>法人':
                        pyperclip.copy(farenName)
                    elif copytext == '>电话':
                        pyperclip.copy(phoneNum)
                    elif copytext == '>彩信内容':
                        caixing = farenName[0] + xingxi
                        pyperclip.copy(caixing)
                    elif copytext == '>字号':
                        if needgongsi == '是':
                            zihao = resPaiban[1]
                            pyperclip.copy(zihao)
                    elif copytext == '>公司名称':
                        if needgongsi == '是':
                            companyName = resPaiban[0]
                            pyperclip.copy(companyName)
                    elif copytext == '无':
                        pass
                    else:
                        pyperclip.copy(copytext)
                    # 执行自动
                    if self.__running.isSet():
                        autoFun(int(x), int(y), event, float(sleepNum))
            # 保存已经执行并且保存从第几行开始 i++对下一个数据进行操作
            value['4'] = i
            value['6'] = i
            config1.write()
            i = i + 1
            if i >= int(jieshu) + 1:
                pyautogui.alert(text='执行完毕', title='提示', button='OK')
                self.stop()

    def pause(self):
        self.__flag.clear()  # 设置为False, 让线程阻塞

    def resume(self):
        self.__flag.set()  # 设置为True, 让线程停止阻塞

    def stop(self):
        self.__flag.set()  # 将线程从暂停状态恢复, 如果已经暂停的话
        self.__running.clear()  # 设置为False

    def killme(self):
        stop_thread(self)


########自动化函数
###自动化执行判断
def autoFun(x, y, event, sleepNum):
    time.sleep(sleepNum)
    if event == '鼠标左键双击':
        pyautogui.doubleClick(x=x, y=y, button="left")
    elif event == '鼠标左键单击':
        pyautogui.click(x=x, y=y)
    elif event == '鼠标右键单击':
        pyautogui.click(x=x, y=y, button='right')
    elif event == '鼠标移动到':
        pyautogui.moveTo(x, y)
    elif '>' in event:
        keywords = str(event).replace('>', '')
        lists = keywords.split('+')
        for i in lists:
            time.sleep(0.2)
            pyautogui.keyDown(i)
        for i in lists:
            time.sleep(0.2)
            pyautogui.keyUp(i)
    else:
        pass

##########
### 公司名称排版
def paiban(path, faren):
    f = open(path, 'r', encoding='utf-8')
    pblist = []
    cnt = 0
    zuida = 0
    resstr = ''
    zihao = ''
    for i in f:
        if cnt == 0:
            zuida = int(i.split('|')[1])
            zihao = int(i.split('|')[3])
            cnt += 1
        else:
            sp = i.replace('\n', '').split('|')
            pblist.append(sp)
    cnt = 0
    flag = False
    for i in pblist:
        if len(i[1].replace(' ', '')) == len(str(faren)):
            zihao = i[2]
            for k in i[1]:
                if k != 'A':
                    resstr += k
                else:
                    resstr += faren[cnt]
                    cnt += 1
            flag = True
            break
        if flag:
            break
    if not flag:
        resstr = faren
        if len(str(resstr))>zuida:
            resstr = resstr[0:zuida]
    f.close()
    return [resstr, int(zihao)]